"""Behavior tests for extract-template CLI flows and related prereq helpers."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, cast

import openpyxl
import pytest

from transitionchecker.cli import extract_template_cli
from transitionchecker.core import Catalogue, CatalogueEntry
from transitionchecker.prereq_engine import (
    build_prerequisite_snapshot,
    classify_prerequisite_clause,
    parse_prerequisite_field,
    salvage_mixed_prerequisite_clause,
    PrerequisiteClauseClassification,
)


BASELINE_UPDATE_HINT = (
    "If you need to update the baseline, run: "
    "extract-template --catalogue-input tests/data/catalogue_prereq_fixture.json "
    "--prereq-snapshot-output tests/data/prereq-snapshot-baseline.json"
)


class _FakeWorkbook:
    def close(self) -> None:
        return None


def test_main_requires_xlsx_or_catalogue_input() -> None:
    code = extract_template_cli.main([])
    assert code == 1


def test_main_returns_1_for_missing_workbook(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    code = extract_template_cli.main([str(missing)])
    assert code == 1


def test_main_writes_snapshot_from_catalogue_input(tmp_path: Path) -> None:
    catalogue_in = tmp_path / "catalogue.json"
    snapshot_out = tmp_path / "snapshot.json"
    catalogue_in.write_text(
        json.dumps(
            [
                {
                    "code": "CEIC1000",
                    "career": "UGRD",
                    "title": "Course",
                    "uoc": 6,
                    "prerequisites": ".",
                }
            ]
        ),
        encoding="utf-8",
    )

    code = extract_template_cli.main(
        [
            "--catalogue-input",
            str(catalogue_in),
            "--prereq-snapshot-output",
            str(snapshot_out),
        ]
    )

    assert code == 0
    snapshot = json.loads(snapshot_out.read_text(encoding="utf-8"))
    assert snapshot["meta"]["source_catalogue"] == str(catalogue_in)
    assert snapshot["entries"][0]["course_code"] == "CEIC1000"


def test_main_success_writes_outputs(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    excel = tmp_path / "mapping.xlsx"
    excel.write_text("placeholder", encoding="utf-8")

    catalogue_out = tmp_path / "plans" / "catalogue.json"
    template_out = tmp_path / "templates" / "template_configs.json"

    def fake_load_workbook(_path: Path, data_only: bool = True) -> _FakeWorkbook:
        return _FakeWorkbook()

    def fake_extract_catalogue(_wb: Any) -> Catalogue:
        return Catalogue(
            [
                CatalogueEntry(
                    code="TEST1001",
                    title="T",
                    career="UGRD",
                    uoc=6,
                    prerequisites=".",
                    level="Level 1",
                )
            ]
        )

    def fake_extract_template_configs(_path: Path) -> dict[str, Any]:
        return {"intakes": {"2026 T1": {"years": []}}}

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(
        extract_template_cli, "extract_catalogue", fake_extract_catalogue
    )
    monkeypatch.setattr(
        extract_template_cli,
        "extract_template_configs_from_workbook",
        fake_extract_template_configs,
    )

    code = extract_template_cli.main(
        [
            str(excel),
            "--catalogue-output",
            str(catalogue_out),
            "--template-output",
            str(template_out),
        ]
    )

    assert code == 0
    assert catalogue_out.is_file()
    assert template_out.is_file()

    catalogue = json.loads(catalogue_out.read_text(encoding="utf-8"))
    templates = json.loads(template_out.read_text(encoding="utf-8"))
    assert catalogue[0]["code"] == "TEST1001"
    assert "intakes" in templates


def test_build_prerequisite_snapshot_is_deterministic() -> None:
    data_dir = Path(__file__).parent / "data"
    catalogue = Catalogue.from_list(
        json.loads(
            (data_dir / "catalogue_prereq_fixture.json").read_text(encoding="utf-8")
        )
    )

    snapshot = build_prerequisite_snapshot(
        catalogue,
        source_catalogue="tests/data/catalogue_prereq_fixture.json",
        generated_at="2026-04-21T00:00:00+00:00",
    )
    expected = json.loads(
        (data_dir / "prereq-snapshot-baseline.json").read_text(encoding="utf-8")
    )

    # generated_at and source_catalogue are intentionally variable and should
    # not block semantic snapshot regression checks.
    for obj in (snapshot, expected):
        meta_obj = obj.get("meta", {})
        if isinstance(meta_obj, dict):
            meta = cast(dict[str, Any], meta_obj)
            meta.pop("generated_at", None)
            meta.pop("source_catalogue", None)

    assert snapshot == expected, BASELINE_UPDATE_HINT


def test_main_writes_prereq_snapshot(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    excel = tmp_path / "mapping.xlsx"
    excel.write_text("placeholder", encoding="utf-8")

    snapshot_out = tmp_path / "snapshots" / "prereq-snapshot.json"
    catalogue_out = tmp_path / "plans" / "catalogue.json"
    template_out = tmp_path / "templates" / "template_configs.json"

    def fake_load_workbook(_path: Path, data_only: bool = True) -> _FakeWorkbook:
        return _FakeWorkbook()

    def fake_extract_catalogue(_wb: Any) -> Catalogue:
        return Catalogue(
            [
                CatalogueEntry(
                    code="TEST1001",
                    title="T",
                    career="UGRD",
                    uoc=6,
                    prerequisites="CEIC1000",
                )
            ]
        )

    def fake_extract_template_configs(_path: Path) -> dict[str, Any]:
        return {"intakes": {"2026 T1": {"years": []}}}

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(
        extract_template_cli, "extract_catalogue", fake_extract_catalogue
    )
    monkeypatch.setattr(
        extract_template_cli,
        "extract_template_configs_from_workbook",
        fake_extract_template_configs,
    )

    code = extract_template_cli.main(
        [
            str(excel),
            "--catalogue-output",
            str(catalogue_out),
            "--template-output",
            str(template_out),
            "--prereq-snapshot-output",
            str(snapshot_out),
        ]
    )

    assert code == 0
    assert snapshot_out.is_file()
    snapshot = json.loads(snapshot_out.read_text(encoding="utf-8"))
    assert snapshot["meta"]["entry_count"] == 1
    assert snapshot["entries"][0]["course_code"] == "TEST1001"
    assert snapshot["entries"][0]["prereq_expr"] == "CEIC1000"
    assert snapshot["entries"][0]["error"] is None


def test_classify_prerequisite_clause_families() -> None:
    classification, families = classify_prerequisite_clause(
        "Must be enrolled in program 4501"
    )
    assert classification is PrerequisiteClauseClassification.IGNORABLE
    assert "program_enrolment" in families

    classification, families = classify_prerequisite_clause(
        "(CEIC2001 OR CEIC2002) and 65+ WAM"
    )
    assert classification is PrerequisiteClauseClassification.MIXED
    assert "wam_mark" in families

    classification, families = classify_prerequisite_clause("CEIC2001 AND CEIC2002")
    assert classification is PrerequisiteClauseClassification.NON_IGNORABLE
    assert families == []

    classification, families = classify_prerequisite_clause(
        "Prerequisite: Enrolment in Accounting Co-op Major (ACCTB13554)"
    )
    assert classification is PrerequisiteClauseClassification.IGNORABLE
    assert "program_enrolment" in families

    classification, families = classify_prerequisite_clause(
        "Students who have previously completed ACCT5906 should not enrol into this course."
    )
    assert classification is PrerequisiteClauseClassification.NON_IGNORABLE
    assert families == []


def test_lint_prerequisites_json_includes_classification(tmp_path: Path) -> None:
    catalogue = Catalogue(
        [
            CatalogueEntry(
                code="A",
                title="A",
                career="UGRD",
                prerequisites="Must be enrolled in program 4501",
            ),
            CatalogueEntry(
                code="B",
                title="B",
                career="UGRD",
                prerequisites="(CEIC2001 OR CEIC2002) and 65+ WAM",
            ),
        ]
    )
    out = tmp_path / "lint.json"

    code = extract_template_cli.lint_prerequisites(catalogue, str(out))

    assert code == 1
    rows = json.loads(out.read_text(encoding="utf-8"))
    by_code = {row["course_code"]: row for row in rows}
    assert (
        by_code["A"]["classification"]
        == PrerequisiteClauseClassification.IGNORABLE.value
    )
    assert "program_enrolment" in by_code["A"]["matched_families"]
    assert (
        by_code["B"]["classification"] == PrerequisiteClauseClassification.MIXED.value
    )
    assert "wam_mark" in by_code["B"]["matched_families"]
    assert by_code["A"]["salvaged"] is False
    assert by_code["A"]["salvaged_expr"] == ""
    assert by_code["A"]["salvage_error"] == ""
    assert by_code["B"]["salvaged"] is True
    assert by_code["B"]["salvaged_expr"]
    assert by_code["B"]["salvage_error"] == ""


def test_salvage_mixed_prerequisite_clause_success() -> None:
    salvaged, salvaged_expr, salvage_error = salvage_mixed_prerequisite_clause(
        "(CEIC2001 OR CEIC2002) and 65+ WAM",
        ["wam_mark"],
    )

    assert salvaged is True
    assert salvaged_expr is not None
    assert salvage_error is None


def test_salvage_mixed_prerequisite_clause_failure() -> None:
    salvaged, salvaged_expr, salvage_error = salvage_mixed_prerequisite_clause(
        "approval from the School",
        ["application_approval"],
    )

    assert salvaged is False
    assert salvaged_expr is None
    assert salvage_error is not None


def test_salvage_mixed_program_enrolment_with_program_title() -> None:
    salvaged, salvaged_expr, salvage_error = salvage_mixed_prerequisite_clause(
        "Prerequisites: EXPT4152 and Enrolment in Program 3897 Applied Exercise Science/Clinical Exercise Physiology",
        ["program_enrolment"],
    )

    assert salvaged is True
    assert salvaged_expr == "EXPT4152"
    assert salvage_error is None


def test_salvage_mixed_language_placement_approval() -> None:
    salvaged, salvaged_expr, salvage_error = salvage_mixed_prerequisite_clause(
        "Prerequisite: ARTS1631 or language placement approval",
        ["application_approval"],
    )

    assert salvaged is True
    assert salvaged_expr == "ARTS1631"
    assert salvage_error is None


def test_salvage_mixed_program_major_clause_with_dangling_comma() -> None:
    salvaged, salvaged_expr, salvage_error = salvage_mixed_prerequisite_clause(
        "Prerequisite: ACTL2102 or (MATH2901 AND MATHE1, MATHM1 or MATHT1 major)",
        ["program_enrolment"],
    )

    assert salvaged is True
    assert salvaged_expr == {"or": ["ACTL2102", "MATH2901"]}
    assert salvage_error is None


def test_parse_prerequisite_field_ignores_or_equivalent() -> None:
    prereq_expr, coreq_expr, error = parse_prerequisite_field(
        "Prerequisite: ACCT5997 or equivalent"
    )

    assert prereq_expr == "ACCT5997"
    assert coreq_expr is None
    assert error is None
